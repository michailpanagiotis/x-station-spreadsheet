#!/usr/bin/env python3
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook, load_workbook

def assert_workbook_sheets_are_same(ws1, ws2, ignore_columns=(1,)):
    for row in ws1.rows:
        for cell in row:
            if cell.column in ignore_columns:
                continue
            other_cell = ws2.cell(row=cell.row, column=cell.column)
            cell_value = cell.value if cell.value is not None else ""
            other_value = other_cell.value if other_cell.value is not None else ""
            if cell_value != other_value:
                raise Exception('difference at cell %s (%s != %s)' % (cell, cell.value, other_cell.value))

def add_references(from_sheet, to_sheet, compare_columns=('B', 'A'), column_refs=(
    ('D', 'B'), ('E', 'C'), ('F', 'D'), ('G', 'E'), ('H', 'F'), ('I', 'G'), ('J', 'H'), ('N', 'I'), ('O', 'J'), ('P', 'K'), ('Q', 'L'), ('R', 'M'), ('S', 'N'),
)):
    target_cell_range = '%s!$%s$2:$P$1001' % (to_sheet.title, compare_columns[1])
    offset_per_column_letter = {x[0]: to_sheet['%s1' % x[1]].column for x in column_refs}

    for row in from_sheet.rows:
        for cell in row:
            if cell.row > 1 and cell.column_letter in offset_per_column_letter:
                offset = offset_per_column_letter[cell.column_letter]
                cell.value = '=IFERROR(VLOOKUP($B%s,%s,%s,0), "")' % (cell.row, target_cell_range, offset)

def fill_rows(sheet, min_row, max_row, color):
    for col in sheet.iter_cols(min_row=min_row, max_row=max_row):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor=color)

def colorize(workbook):
    bold = Font(bold=True)
    sheet = workbook["Template configuration"]
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 16
    for col in sheet.iter_cols(min_row=1, max_col=1):
        for cell in col:
            cell.font = bold

    fill_rows(sheet, 4, 17, "ccd2ff")
    fill_rows(sheet, 18, 31, "b9c4ff")

    # Delay
    fill_rows(sheet, 34, 39, "ccd2ff")
    fill_rows(sheet, 76, 83, "ccd2ff")

    # Reverb
    fill_rows(sheet, 40, 44, "ffc1f5")
    fill_rows(sheet, 84, 88, "ffc1f5")

    # Chorus
    fill_rows(sheet, 45, 51, "d5ffd9")
    fill_rows(sheet, 89, 97, "d5ffd9")

    # Compress
    fill_rows(sheet, 52, 62, "ffc9c8")
    fill_rows(sheet, 98, 108, "ffc9c8")

    # Distortion
    fill_rows(sheet, 63, 66, "d6ffff")
    fill_rows(sheet, 109, 113, "d6ffff")

    # EQ
    fill_rows(sheet, 67, 69, "ffffc1")
    fill_rows(sheet, 114, 116, "ffffc1")

    # Keyboard zones
    fill_rows(sheet, 124, 124, "ccd2ff")
    fill_rows(sheet, 125, 131, "b9c4ff")
    fill_rows(sheet, 132, 138, "ccd2ff")
    fill_rows(sheet, 139, 145, "b9c4ff")
    fill_rows(sheet, 146, 152, "ccd2ff")


    sheet = workbook["Templates"]
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 6
    sheet.column_dimensions["C"].width = 6
    sheet.column_dimensions["D"].width = 6
    sheet.column_dimensions["E"].width = 13
    sheet.column_dimensions["F"].width = 6
    sheet.column_dimensions["J"].width = 6
    sheet.column_dimensions["K"].width = 6
    sheet.column_dimensions["L"].width = 37
    sheet.column_dimensions["M"].width = 6
    sheet.column_dimensions["N"].width = 6
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold

    sheet = workbook["Controls"]
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 6
    sheet.column_dimensions["E"].width = 6
    sheet.column_dimensions["F"].width = 6
    sheet.column_dimensions["G"].width = 13
    sheet.column_dimensions["H"].width = 6
    sheet.column_dimensions["K"].width = 4
    sheet.column_dimensions["O"].width = 6
    sheet.column_dimensions["P"].width = 6
    sheet.column_dimensions["Q"].width = 37
    sheet.column_dimensions["R"].width = 6
    sheet.column_dimensions["S"].width = 6
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold

def create_sheet(workbook, sheet_name, values_table):
    workbook.create_sheet(sheet_name)
    wst = workbook[sheet_name]
    for row_idx, row_values in enumerate(values_table):
        for col_idx, value in enumerate(row_values):
            wst.cell(row=row_idx + 1, column=col_idx + 1, value=value)
