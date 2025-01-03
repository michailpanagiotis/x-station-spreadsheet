#!/usr/bin/env python3
import json
from openpyxl import Workbook, load_workbook
from fields import define_field, NumericArray, NumericValue, SelectValue, BitMap, StringValue, ZeroPadding, FieldSet
from xlsx_utils import assert_workbook_sheets_are_same, create_sheet, add_references, colorize

NUM_TEMPLATE_CONFIGURATION_BYTES = 396
MESSAGE_START=b'\xf0\x00 )\x02\x00\x7f\x00\x00'
MESSAGE_END=b'\x124\xf7'
CONTROL_LINE_SIZE = 52

with open('x-station-indices.json', 'r') as f:
    indices = json.loads(f.read())

def get_control_section(idx):
    return indices[idx]['section']

def get_control_physical(idx):
    return indices[idx].get('physical', None)

def get_control_label(idx):
    if 'selector' not in indices[idx]:
        return indices[idx]['legend']
    return '%s (%s)' % (indices[idx]['legend'], indices[idx]['selector'])

def get_control_legend(idx):
    selector = '(%s)' % indices[idx]['selector'] if 'selector' in indices[idx] else ''
    group = '%s%s' % (indices[idx]['section'], selector)
    legend = '%s>%s' % (group, indices[idx]['legend'])
    return legend.strip()

def _assert_bytes_are_same(bytes1, bytes2):
    for idx, byte in enumerate(bytes1):
        if byte != bytes2[idx]:
            raise Exception('bad serializing %s %s' % (byte, bytes2[idx]))

Sysex = define_field(NumericArray, num_bytes=18, name="Sysex")
ControlName = define_field(StringValue, num_bytes=16, name="Name", aliases=['Control name'])
TemplateName = define_field(StringValue, num_bytes=16)
ManufacturerName = define_field(StringValue, num_bytes=13)
Pad1 = define_field(ZeroPadding, num_bytes=1, name="Zeros")
Pad2 = define_field(ZeroPadding, num_bytes=2, name="Zeros")
Pad3 = define_field(ZeroPadding, num_bytes=3, name="Zeros")
Pad4 = define_field(ZeroPadding, num_bytes=4, name="Zeros")
Pad5 = define_field(ZeroPadding, num_bytes=5, name="Zeros")
Pad6 = define_field(ZeroPadding, num_bytes=6, name="Zeros")
Pad7 = define_field(ZeroPadding, num_bytes=7, name="Zeros")
Pad8 = define_field(ZeroPadding, num_bytes=8, name="Zeros")
Pad166 = define_field(ZeroPadding, num_bytes=166, name="Zeros")
Unknown = SelectValue

CONTROL_FIELDS = [
    ControlName,
    define_field(NumericValue, name="Type", aliases=['Control Type']),
    define_field(NumericValue, name="Low", aliases=['Template', 'Velocity', 'MMC Command']),
    define_field(NumericValue, name="High"),
    define_field(BitMap, stuffed=(['Ports', 4], ['Button', 4])),
    define_field(NumericValue, name='Pot', aliases=['Pot / Slider Control Type']),
    define_field(NumericValue, name='Display', aliases=['Display type']),
    define_field(NumericValue, name='MSBank', aliases=['NRPN MSBank Num']),
    define_field(NumericValue, name='CC', aliases=['Note']),
    define_field(NumericValue, name='Channel', aliases=['Channel', 'Device id']),
    define_field(NumericValue, name='Default', aliases=['Template', 'Velocity', 'MMC Command']),
    define_field(NumericValue, name='HasSysex'),
    define_field(NumericValue, name='N/A 2'),
    define_field(NumericValue, name='N/A 3'),
    Sysex,
    define_field(NumericValue, name='Step'),
    Pad4,
]

CONTROL_TEMPLATE_FIELDS = [
    define_field(NumericValue, name="Type", aliases=['Control Type']),
    define_field(NumericValue, name="Low", aliases=['Template', 'Velocity', 'MMC Command']),
    define_field(NumericValue, name="High"),
    define_field(BitMap, stuffed=(['Ports', 4], ['Button', 4])),
    define_field(NumericValue, name='Pot', aliases=['Pot / Slider Control Type']),
    define_field(NumericValue, name='Display', aliases=['Display type']),
    define_field(NumericValue, name='MSBank', aliases=['NRPN MSBank Num']),
    define_field(NumericValue, name='HasSysex'),
    define_field(NumericValue, name='N/A 2'),
    define_field(NumericValue, name='N/A 3'),
    Sysex,
    define_field(NumericValue, name='Step'),
    Pad4,
]

CONTROL_TEMPLATE_PHYSICAL_FIELDS = [
    define_field(NumericValue, name='Display', aliases=['Display type']),
]

TEMPLATE_FIELDS = [
    define_field(Unknown, name='Unknown', valid_values=[8, 17, 19, 24, 25]),
    define_field(Unknown, name='Unknown', valid_values=[1, 2, 3, 10]),
    Pad1,
    define_field(NumericValue, name='Template index'),
    define_field(TemplateName, name='Name'),
    define_field(NumericValue, name='Internal index'),
    define_field(ManufacturerName, name='Manufacturer'),
    define_field(SelectValue, name='Channel', valid_values=[0, 16]),
    define_field(SelectValue, name='Midi port | Keyb MIDI Chan', valid_values=[0, 16, 48, 53, 54, 112]),
    define_field(NumericValue, name='Bank'),
    define_field(NumericValue, name='Program'),
    define_field(SelectValue, name='Velocity curve', valid_values=[0, 1, 2, 3]),
    define_field(NumericValue, name='Octave setting'),
    define_field(SelectValue, name='Aftertouch | Auto Snapshot | Not Synth', valid_values=[0, 1, 2, 3, 4, 5, 6, 7]),
    define_field(NumericValue, name='Override MIDI Channel'),
    define_field(SelectValue, name='Touchpad X Type', valid_values=[0, 1, 2]),
    define_field(SelectValue, name='Touchpad Y Type', valid_values=[0, 1, 2]),
    define_field(SelectValue, name='Stereo', valid_values=[0, 1, 2]),
    Pad1,
    define_field(NumericValue, name='Input 1 - Gain'),
    define_field(NumericValue, name='Input 1 - Pan'),
    Pad1,
    define_field(NumericValue, name='Input 1 - Bypass Effects'),
    define_field(NumericValue, name='Input 2 - Gain'),
    define_field(NumericValue, name='Input 2 - Pan'),
    Pad1,
    define_field(NumericValue, name='Input 2 - Bypass Effects'),
    define_field(NumericValue, name='Stereo - Gain'),
    define_field(NumericValue, name='Stereo - Width'),
    Pad1,
    define_field(NumericValue, name='Stereo - Bypass Effects'),
    Pad1,
    define_field(Unknown, name='Unknown', valid_values=[0,1,64]),
    define_field(NumericValue, name='Input 1 - Delay - Level'),
    define_field(NumericValue, name='Input 1 - Delay - Decay time'),
    define_field(NumericValue, name='Input 1 - Delay - Feedback'),
    define_field(NumericValue, name='Input 1 - Delay - Stereo Width'),
    define_field(NumericValue, name='Input 1 - Delay - L/R Ratio'),
    Pad5,
    define_field(NumericValue, name='Input 1 - Reverb - Level'),
    define_field(NumericValue, name='Input 1 - Reverb - Type'),
    define_field(NumericValue, name='Input 1 - Reverb - Decay'),
    define_field(Unknown, name='Unknown', valid_values=[0,64]),
    Pad6,
    define_field(NumericValue, name='Input 1 - Chorus - Level'),
    define_field(NumericValue, name='Input 1 - Chorus - Type'),
    define_field(NumericValue, name='Input 1 - Chorus - Rate'),
    define_field(NumericValue, name='Input 1 - Chorus - Mod Depth'),
    define_field(NumericValue, name='Input 1 - Chorus - Mod Centre'),
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    Pad4,
    define_field(NumericValue, name='Input 1 - Compress - Ratio'),
    define_field(NumericValue, name='Input 1 - Compress - Threshold'),
    define_field(NumericValue, name='Input 1 - Compress - Attack'),
    define_field(NumericValue, name='Input 1 - Compress - Release'),
    define_field(Unknown, name='Unknown', valid_values=[0, 3, 32]),
    Pad1,
    define_field(NumericValue, name='Input 1 - Compress - Auto Gain'),
    define_field(Unknown, name='Unknown', valid_values=[0, 3, 64]),
    Pad1,
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    Pad2,
    define_field(NumericValue, name='Input 1 - Distortion - Level'),
    define_field(NumericValue, name='Input 1 - Distortion - Compensate'),
    define_field(NumericValue, name='Input 1 - Distortion - Output Level'),
    Pad3,
    define_field(NumericValue, name='Input 1 - EQ - Low'),
    define_field(NumericValue, name='Input 1 - EQ - High'),
    define_field(NumericValue, name='Input 1 - EQ - Mid'),
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    define_field(Unknown, name='Unknown', valid_values=[0, 64, 127]),
    define_field(Unknown, name='Unknown', valid_values=[0, 2, 64]),
    Pad6,
    define_field(Unknown, name='Unknown', valid_values=[0, 4, 64]),
    Pad3,
    define_field(NumericValue, name='Input 2 - Delay - Level'),
    define_field(NumericValue, name='Input 2 - Delay - Decay time'),
    define_field(NumericValue, name='Input 2 - Delay - Feedback'),
    define_field(NumericValue, name='Input 2 - Delay - Stereo Width'),
    define_field(NumericValue, name='Input 2 - Delay - L/R Ratio'),
    Pad1,
    define_field(Unknown, name='Unknown', valid_values=[0, 64, 78]),
    Pad3,
    define_field(NumericValue, name='Input 2 - Reverb - Level'),
    define_field(NumericValue, name='Input 2 - Reverb - Type'),
    define_field(NumericValue, name='Input 2 - Reverb - Decay'),
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    Pad6,
    define_field(NumericValue, name='Input 2 - Chorus - Level'),
    define_field(NumericValue, name='Input 2 - Chorus - Type'),
    define_field(NumericValue, name='Input 2 - Chorus - Rate'),
    define_field(NumericValue, name='Input 2 - Chorus - Mod Depth'),
    define_field(NumericValue, name='Input 2 - Chorus - Mod Centre'),
    define_field(Unknown, name='Unknown', valid_values=[0, 1, 64, 74]),
    Pad1,
    define_field(Unknown, name='Unknown', valid_values=[0, 2, 64]),
    Pad2,
    define_field(NumericValue, name='Input 2 - Compress - Ratio'),
    define_field(NumericValue, name='Input 2 - Compress - Threshold'),
    define_field(NumericValue, name='Input 2 - Compress - Attack'),
    define_field(NumericValue, name='Input 2 - Compress - Release'),
    define_field(Unknown, name='Unknown', valid_values=[0,32,64]),
    Pad1,
    define_field(NumericValue, name='Input 2 - Compress - Auto Gain'),
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    Pad1,
    define_field(Unknown, name='Unknown', valid_values=[0, 64]),
    Pad2,
    define_field(NumericValue, name='Input 2 - Distortion - Level'),
    define_field(NumericValue, name='Input 2 - Distortion - Compensate'),
    define_field(NumericValue, name='Input 2 - Distortion - Output Level'),
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 64, 65]),
    Pad2,
    define_field(NumericValue, name='Input 2 - EQ - Low'),
    define_field(NumericValue, name='Input 2 - EQ - High'),
    define_field(NumericValue, name='Input 2 - EQ - Mid'),
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 64]),
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 64]),
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 64]),
    Pad6,
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 64]),
    Pad3,
    define_field(Unknown, name='Unknown Flag', valid_values=[0, 1]),
    define_field(SelectValue, name='Enable Keyboard Zones', valid_values=[0, 1, 64]),
    define_field(NumericValue, name='Zone 1 - Midi Channel | Midi ports'),
    define_field(NumericValue, name='Zone 1 - Velocity Offset'),
    define_field(NumericValue, name='Zone 1 - Bottom Note'),
    define_field(NumericValue, name='Zone 1 - Top Note'),
    define_field(NumericValue, name='Zone 1 - Transpose'),
    define_field(NumericValue, name='Zone 1 - Aftertouch | Pitch Bend | Mod Wheel'),
    Pad4,
    define_field(NumericValue, name='Zone 2 - Midi Channel | Midi ports'),
    define_field(NumericValue, name='Zone 2 - Velocity Offset'),
    define_field(NumericValue, name='Zone 2 - Bottom Note'),
    define_field(NumericValue, name='Zone 2 - Top Note'),
    define_field(NumericValue, name='Zone 2 - Transpose'),
    define_field(NumericValue, name='Zone 2 - Aftertouch | Pitch Bend | Mod Wheel'),
    Pad4,
    define_field(NumericValue, name='Zone 3 - Midi Channel | Midi ports'),
    define_field(NumericValue, name='Zone 3 - Velocity Offset'),
    define_field(NumericValue, name='Zone 3 - Bottom Note'),
    define_field(NumericValue, name='Zone 3 - Top Note'),
    define_field(NumericValue, name='Zone 3 - Transpose'),
    define_field(NumericValue, name='Zone 3 - Aftertouch | Pitch Bend | Mod Wheel'),
    Pad4,
    define_field(NumericValue, name='Zone 4 - Midi Channel | Midi ports'),
    define_field(NumericValue, name='Zone 4 - Velocity Offset'),
    define_field(NumericValue, name='Zone 4 - Bottom Note'),
    define_field(NumericValue, name='Zone 4 - Top Note'),
    define_field(NumericValue, name='Zone 4 - Transpose'),
    define_field(NumericValue, name='Zone 4 - Aftertouch | Pitch Bend | Mod Wheel'),
    Pad4,
    Pad166,
]


REALEARN_TARGET_DUMMY = {
    "type": 53,
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "useProject": False,
    "moveView": False,
    "seekPlay": False,
    "oscArgIndex": 0,
    "mouseAction": { "kind": "MoveTo", "axis": "X" },
    "takeMappingSnapshot": { "kind": "LastLoaded" }
}

REALEARN_BUTTON_DESIGN = {
    "background": { "kind": "Color" },
    "foreground": { "kind": "None" },
    "static_text": ""
}

REALEARN_CONTROL_SOURCE_COMMON = {
    "isRegistered": False,
    "is14Bit": False,
    "oscArgIndex": 0,
    "buttonIndex": 0,
    "buttonDesign": REALEARN_BUTTON_DESIGN,
}

REALEARN_CONTROL_TARGET_COMMON = {
    "category": "virtual",
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "learnable": False,
    "mouseAction": {
      "kind": "MoveTo",
      "axis": "X"
    },
    "pollForFeedback": False,
    "takeMappingSnapshot": {
      "kind": "ById",
      "id": ""
    }
}

KNOWN_TEMPLATES = [FieldSet.from_csv(CONTROL_TEMPLATE_FIELDS, csv, name=name) for name, csv in ({
    "Trigger": "1,127,0,01110000,4,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Toggle": "1,0,127,01111000,4,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Momentary": "1,0,127,01110100,0,0,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Continuous": "1,0,127,01110000,0,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "ContinuousPickup": "1,0,127,01110000,9,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Unary": "1,63,64,01110000,0,65,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Jog": "1,63,65,01110000,0,65,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Pitch": "10,0,127,01110000,0,65,0,0,0,0,000000000000000000000000000000000000,0,4",
    "-": "0,0,127,01110000,0,64,0,0,0,0,000000000000000000000000000000000000,0,4",
}).items()]

def split_sysex(sysex):
    if sysex[:len(MESSAGE_START)] != MESSAGE_START:
        raise Exception('bad header')

    if sysex[-len(MESSAGE_END):] != MESSAGE_END:
        raise Exception('bad footer')

    template_configuration_bytes = bytearray(
        sysex[len(MESSAGE_START):len(MESSAGE_START) + NUM_TEMPLATE_CONFIGURATION_BYTES],
    ) # 396 bytes

    controls_bytes = bytearray(
        sysex[len(MESSAGE_START) + len(template_configuration_bytes):-len(MESSAGE_END)] )
    # 52 x 149 = 7748 bytes
    return template_configuration_bytes, controls_bytes

def extract_templates(controls, definition=CONTROL_TEMPLATE_FIELDS, known=KNOWN_TEMPLATES):
    permutations = {t.get_subset(definition) for t in controls}
    templates = []
    for idx, template in enumerate(permutations):
        if len(template.fields) != len(definition):
            raise Exception('bad template')
        name = 'template%s' % idx
        match = next((x for x in known if x == template), None)
        template.add_labels(Name=match.name if match else name)
        templates.append(template)
    templates.sort(key=lambda t: t.bytes)
    return templates


class SingleControl(FieldSet):
    @classmethod
    def from_bytes(cls, *args, **kwargs):
        return super(SingleControl, cls).from_bytes(CONTROL_FIELDS, *args, **kwargs)

    @classmethod
    def from_values(cls, *args, **kwargs):
        return super(SingleControl, cls).from_values(CONTROL_FIELDS, *args, **kwargs)

    def __init__(self, fields, index):
        super().__init__(fields)
        self.index = index

    def get_template(self):
        return self.get_subset(CONTROL_TEMPLATE_FIELDS)

    def get_references(self, other_fieldset_definition=CONTROL_TEMPLATE_FIELDS):
        other_fields = self.get_subset(CONTROL_TEMPLATE_FIELDS)
        references = []
        for idx, field in enumerate(self.fields):
            other_idx = other_fields.find_index(field.name)
            if other_idx is not None:
                references.append((idx, other_idx))
        return references

    def to_realearn_dict(self, idx):
        id = '%s. %s' % (idx, get_control_legend(idx))
        control = {
          "id": id,
          "name": get_control_label(idx),
          "groupId": get_control_section(idx),
          "source": {
            **REALEARN_CONTROL_SOURCE_COMMON,
            "character": 1 if get_control_physical(idx) == 'Button' else 0,
            "channel": int(str(self['Channel'])),
            "number": int(str(self['CC'])),
          },
          "mode": {
            "maxStepSize": 0.05,
            "minStepFactor": 1,
            "maxStepFactor": 5,
            "takeoverMode": "pickup-tolerant"
          },
          "target": {
                **REALEARN_CONTROL_TARGET_COMMON,
                "controlElementType": "Button" if get_control_physical(idx) == 'Button' else None,
                "controlElementIndex": "%s-%s" % (get_control_section(idx), get_control_label(idx)),
          },
          "feedbackIsEnabled": False,
          "visibleInProjection": False
        }
        return control

class Template():
    def __init__(self, header_fields, controls, sysex=None):
        self.header_fields = header_fields
        self.controls = controls
        self.templates = extract_templates(self.controls)
        for idx, control in enumerate(self.controls):
            found_template = next((x for x in self.templates if x == control.get_subset(CONTROL_TEMPLATE_FIELDS)), None)

            if not found_template:
                raise Exception('unknown template')

            control.add_labels(Legend=get_control_legend(idx), Template=found_template.get_label('Name'))
        self.sysex = sysex

    @property
    def bytes(self):
        bytes = bytearray(MESSAGE_START)
        for field in self.header_fields.fields:
            bytes.extend(field.bytes)
        for control in self.controls:
            bytes.extend(control.bytes)
        bytes.extend(MESSAGE_END)
        return bytes

    def __getitem__(self, key):
        field = next((x for x in self.header_fields if x.name == key), None)
        if field is None:
            raise KeyError
        return field

    def __str__(self):
       return '\n'.join([str(x) for x in self.controls])

    def __len__(self):
        return len(self.bytes)

    def __eq__(self, other):
        if not isinstance(other, Template):
            # don't attempt to compare against unrelated types
            return NotImplemented

        if len(self) != len(other):
            return False

        for idx, byte in enumerate(self.bytes):
            if byte != other.bytes[idx]:
                return False
        return True

    @property
    def name(self):
        name = next(str(x) for x in self.header_fields if x.name == 'Name')
        manufacturer = next(str(x) for x in self.header_fields if x.name == 'Manufacturer')
        return ('%s %s' % (manufacturer.strip(), name.strip())).strip()

    @property
    def unknowns(self):
        return [x for x in self.controls if x.section == '']

    @classmethod
    def from_sysex(cls, filename):
        with open(filename, "rb") as f:
            file_contents = f.read()

        template_configuration_bytes, controls_bytes = split_sysex(file_contents)

        instance = cls(
            header_fields=FieldSet.from_bytes(TEMPLATE_FIELDS, template_configuration_bytes),
            controls = [
                SingleControl.from_bytes(bytearray(controls_bytes[i:i + CONTROL_LINE_SIZE]), index=idx)
                for idx, i in enumerate(range(0, len(controls_bytes), CONTROL_LINE_SIZE))
            ],
            sysex = file_contents
        )

        _assert_bytes_are_same(file_contents, instance.bytes)
        return instance

    def to_sysex(self, file):
        with open(file, "wb") as f:
            f.write(self.bytes)

    @classmethod
    def _from_workbook(cls, workbook):
        ws = workbook['Template configuration']
        header_fields = FieldSet.from_values(TEMPLATE_FIELDS, [row[1].value for row in ws.rows])

        wst = workbook['Templates']
        templates = {}
        for idx, row in enumerate(wst.rows):
            if idx == 0:
                continue
            (name, *template_values) = (c.value for c in row)
            templates[name] = FieldSet.from_values(CONTROL_TEMPLATE_FIELDS, template_values, name=name)

        ws = workbook['Controls']
        controls = []
        for idx, row in enumerate(ws.rows):
            if idx == 0:
                continue
            (legend, template_name, *values) = (c.value for c in row)
            template = templates[template_name]
            flat_values = template.dereference(values)
            controls.append(SingleControl.from_values(flat_values, index=idx - 1))
        return cls(header_fields, controls)

    @classmethod
    def from_spreadsheet(cls, filename):
        wb = load_workbook(filename=filename)
        template = cls._from_workbook(wb)
        parsed = template._to_workbook()
        assert_workbook_sheets_are_same(wb['Template configuration'], parsed['Template configuration'])
        assert_workbook_sheets_are_same(wb['Templates'], parsed['Templates'])
        assert_workbook_sheets_are_same(wb['Controls'], parsed['Controls'], ignore_columns=(1, 2))
        return template

    def _to_workbook(self):
        wb = Workbook()
        create_sheet(
            wb,
            "Template configuration",
            list(zip(self.header_fields.get_headers(), self.header_fields.get_values())),
        )

        templates_table = FieldSet.get_table(self.templates, with_labels=True)
        create_sheet(wb, "Templates", templates_table)

        controls_table = FieldSet.get_table(self.controls, with_labels=True)
        create_sheet(wb, "Controls", controls_table)

        del wb['Sheet']

        add_references(wb['Controls'], wb['Templates'])
        colorize(wb)

        return wb

    def to_spreadsheet(self, filename):
        wb = self._to_workbook()
        wb.save(filename)
        stored = Template._from_workbook(wb)
        wb.close()

        if self.bytes != stored.bytes:
            raise Exception('template could not be stored properly')

    def to_json(self, filename):
        id = self.name
        groups = {get_control_section(idx) for idx, c in enumerate(self.controls) if get_control_section(idx) != ''}
        main = {
          "version": "2.16.14",
          "name": "x-station",
          "defaultGroup": {},
          "groups": [{"id": s, "name": s} for s in groups],
          "mappings": [c.to_realearn_dict(idx) for idx, c in enumerate(self.controls) if get_control_physical(idx) is not None],
        }
        with open(filename, "w") as f:
            f.write(json.dumps(main, indent=2))

    def __get_control_legend(self, idx):
        selector = '(%s)' % indices[idx]['selector'] if 'selector' in indices[idx] else ''
        group = '%s%s' % (indices[idx]['section'], selector)
        legend = '%s>%s' % (group, indices[idx]['legend'])
        return legend.strip()

    def get_control_values(self):
        return [control.get_values() for control in self.controls]

    def to_sql(self):
        values = [
            [
                get_control_legend(idx),
                next((x.name for x in self.templates if x == control.get_subset(CONTROL_TEMPLATE_FIELDS)), None),
            ] +  control.get_values()
            for idx, control in enumerate(self.controls)
        ]

        for value in values:
            print(value)

        # for control in self.controls:
        #     refs = control.get_references()
        #     print(refs)

        return values

    def control_permutations(self, field_name):
        return {str(c[field_name]) for c in self.controls}

    def diff(self, other):
        template_diffs = []
        field_diffs = []

        other_headers = other.header_fields
        for idx, field in enumerate(self.header_fields):
            if field != other_headers[idx]:
                template_diffs.append([idx, field, str(field), str(other_headers[idx])])

        other_controls = other.controls
        for idx, control in enumerate(self.controls):
            for field_idx, field in enumerate(control.fields):
                other_field = other_controls[idx].fields[field_idx]
                if field != other_field:
                    print(control)
                    print(other_controls[idx])
                    field_diffs.append([idx, field_idx, field.name, str(field), str(other_field)])

        if len(template_diffs) > 0:
            print('TEMPLATE DIFFS')
            for diff in template_diffs:
                print("\tRow %s ('%s')  '%s' -> %s" % (diff[0] + 1, diff[1].name, diff[2], diff[3]))
        if len(field_diffs) > 0:
            print('FIELD DIFFS', field_diffs)
