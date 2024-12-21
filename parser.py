#!/usr/bin/env python3
import argparse
import json
import string
import sys
from openpyxl import Workbook, load_workbook
from pathlib import Path

with open('x-station-indices.json', 'r') as f:
    indices = json.loads(f.read())

REALEARN_TARGET_DUMMY = {
    "type": 53,
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "useProject": False,
    "moveView": False,
    "seekPlay": False,
    "oscArgIndex": 0,
    "mouseAction": { "kind": "MoveTo", "axis": "X" },
    "takeMappingSnapshot": { "kind": "LastLoaded" }
}

REALEARN_BUTTON_DESIGN = {
    "background": { "kind": "Color" },
    "foreground": { "kind": "None" },
    "static_text": ""
}

REALEARN_CONTROL_SOURCE_COMMON = {
    "isRegistered": False,
    "is14Bit": False,
    "oscArgIndex": 0,
    "buttonIndex": 0,
    "buttonDesign": REALEARN_BUTTON_DESIGN,
}

REALEARN_CONTROL_TARGET_COMMON = {
    "category": "virtual",
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "learnable": False,
    "mouseAction": {
      "kind": "MoveTo",
      "axis": "X"
    },
    "pollForFeedback": False,
    "takeMappingSnapshot": {
      "kind": "ById",
      "id": ""
    }
}

class RawBytes():
    DEFAULTS = {
        "name": None,
        "ms_name": None,
        "ls_name": None,
        "aliases": (),
        "valid_values": (),
    }

    @classmethod
    def _pop_from(cls, other_bytes, *args, **kwargs):
        num_bytes = cls.NUM_BYTES

        instance = cls(other_bytes[:num_bytes], *args, **kwargs)

        if isinstance(other_bytes, bytearray):
            for _ in range(len(instance)):
                other_bytes.pop(0)

        return instance

    @classmethod
    def parse_string(cls, string):
        raise Exception('not implemented for %s' % cls.__name__)

    def __init__(self, value, name=None, *args, **kwargs):
        if value is None:
            raise Exception('value is required')
        if isinstance(value, bytearray):
            if len(value) != type(self).NUM_BYTES:
                raise Exception('bad number of bytes')
            self.bytes = bytearray(value)
        else:
            self.bytes = type(self).parse_string(value)
            if not isinstance(self.bytes, bytearray):
                raise Exception('parsing should return a bytearray but got %s %s' % (self.bytes, type(self.bytes)))

        self.name = name if name is not None else type(self).DEFAULTS['name']
        for argname in ['ms_name', 'ls_name', 'valid_values', 'aliases']:
            setattr(self, argname, kwargs[argname] if argname in kwargs else type(self).DEFAULTS[argname])

        if self.ms_name and self.ls_name:
            self.name = '%s|%s' % (self.ms_name, self.ls_name)

        if len(self.valid_values) > 0:
            for byte in self.bytes:
               if byte not in self.valid_values:
                    raise Exception('unsupported option %s' % byte)


    def __str__(self):
        return ''.join([hex(x) for x in self.bytes])

    def __len__(self):
        return len(self.bytes)

    def __eq__(self, other):
        if len(self) != len(other):
            return False

        for idx, byte in enumerate(self.bytes):
            if byte != other.bytes[idx]:
                return False
        return True


class SingleByte(RawBytes):
    NUM_BYTES = 1

    @classmethod
    def parse_string(cls, string):
        return bytearray([int(string)])

class NumericValue(SingleByte):
    def __str__(self):
        return str(self.bytes[0])

class NumericArray(RawBytes):
    @classmethod
    def parse_string(cls, string):
        return bytearray(bytes.fromhex(string))

    def __str__(self):
        return self.bytes.hex()


class SelectValue(NumericValue):
    def __init__(self, *args, **kwargs):
        if len(type(self).DEFAULTS["valid_values"]) == 0:
            raise Exception("select with no valid values")
        super().__init__(*args, **kwargs)

class StringValue(RawBytes):
    @classmethod
    def parse_string(cls, string):
        return bytearray(string.encode('ascii').ljust(cls.NUM_BYTES, b' '))

    def __str__(self):
        return ''.join([chr(x) for x in self.bytes if chr(x) in string.printable]).rstrip()

class ZeroPadding(RawBytes):
    DEFAULTS = {
        **RawBytes.DEFAULTS,
        "valid_values": (0,),
    }

    @classmethod
    def parse_string(cls, string):
        return bytearray(b'\x00') * cls.NUM_BYTES

    def __str__(self):
        return str(len(self))

class BitMap(SingleByte):
    @classmethod
    def parse_string(cls, string):
        return bytearray([int(string, 16)])

    def __repr__(self):
        formatted = "{:08b}".format(self.bytes[0])
        return '<%s:%s|%s:%s>' % (self.ms_name, formatted[:4], self.ls_name, formatted[4:])

    def __str__(self):
        return hex(self.bytes[0])

def define_field(base_cls, num_bytes=None, **defaults):
    name = defaults["name"] if "name" in defaults else ""
    return type(name, (base_cls,), {
        "NUM_BYTES": num_bytes if num_bytes is not None else base_cls.NUM_BYTES,
        "DEFAULTS": {**base_cls.DEFAULTS, **defaults},
    })

Sysex = define_field(NumericArray, num_bytes=18, name="Sysex")
ControlName = define_field(StringValue, num_bytes=16, name="Name", aliases=['Control name'])
TemplateName = define_field(StringValue, num_bytes=16)
ManufacturerName = define_field(StringValue, num_bytes=13)
Pad1 = define_field(ZeroPadding, num_bytes=1, name="Zeros")
Pad2 = define_field(ZeroPadding, num_bytes=2, name="Zeros")
Pad3 = define_field(ZeroPadding, num_bytes=3, name="Zeros")
Pad4 = define_field(ZeroPadding, num_bytes=4, name="Zeros")
Pad5 = define_field(ZeroPadding, num_bytes=5, name="Zeros")
Pad6 = define_field(ZeroPadding, num_bytes=6, name="Zeros")
Pad7 = define_field(ZeroPadding, num_bytes=7, name="Zeros")
Pad8 = define_field(ZeroPadding, num_bytes=8, name="Zeros")
Pad170 = define_field(ZeroPadding, num_bytes=170, name="Zeros")

class SingleControl():
    FIELD_TYPES = [
        ControlName,
        define_field(NumericValue, name="Type", aliases=['Control Type']),
        define_field(NumericValue, name="Low", aliases=['Template', 'Velocity', 'MMC Command']),
        define_field(NumericValue, name="High"),
        define_field(BitMap, ms_name='Ports', ls_name='Button'),
        define_field(NumericValue, name='Pot', aliases=['Pot / Slider Control Type']),
        define_field(NumericValue, name='Display', aliases=['Display type']),
        define_field(NumericValue, name='MSBank', aliases=['NRPN MSBank Num']),
        define_field(NumericValue, name='CC', aliases=['Note']),
        define_field(NumericValue, name='Ch', aliases=['Channel', 'Device id']),
        define_field(NumericValue, name='Default', aliases=['Template', 'Velocity', 'MMC Command']),
        define_field(NumericValue, name='N/A 1'),
        define_field(NumericValue, name='N/A 2'),
        define_field(NumericValue, name='N/A 3'),
        Sysex,
        define_field(NumericValue, name='Step'),
        Pad4,
    ]

    def __init__(self, idx, fields):
        name = next(x.bytes for x in fields if x.name == 'Name')
        self.index = idx
        self.full_name = ''.join([chr(x) for x in name])
        self.name = self.full_name.strip()
        self.fields = fields

    def __str__(self):
        return '%s' % (' '.join([str(x) for x in self.fields]))

    @property
    def bytes(self):
        bytes = bytearray()
        for field in self.fields:
            bytes.extend(field.bytes)
        return bytes

    @property
    def section(self):
        return indices[self.index]['section']

    @property
    def group(self):
        selector = '(%s)' % indices[self.index]['selector'] if 'selector' in indices[self.index] else ''
        group = '%s%s' % (indices[self.index]['section'], selector)
        return group

    @property
    def label(self):
        if 'selector' not in indices[self.index]:
            return indices[self.index]['legend']
        return '%s (%s)' % (indices[self.index]['legend'], indices[self.index]['selector'])

    @property
    def physical(self):
        return indices[self.index].get('physical', None)

    @property
    def legend(self):
        legend = '%s>%s' % (self.group, indices[self.index]['legend'])
        return legend

    def __getitem__(self, key):
        field = next((x for x in self.fields if x.name == key), None)
        if field is None:
            raise KeyError
        return field

    def __len__(self):
        return len(self.bytes)

    def dict(self):
        j = { "legend": self.legend.strip() }
        for f in self.fields:
            j[f.name] = str(f)
        return j

    def csv(self):
        return '%s,%s' % (self.legend.strip(), ','.join([str(x) for x in self.fields]))

    @classmethod
    def from_sysex(cls, idx, cmd):
        if isinstance(cmd, bytes):
            if len(cmd) != 52:
                raise Exception('bad length')

            cmd = bytearray(cmd)

        fields = [ct._pop_from(cmd) for ct in cls.FIELD_TYPES]

        if len(cmd) != 0:
            raise Exception('non parsed fields')

        return cls(idx, fields)

    @classmethod
    def from_spreadsheet(cls, idx, row):
        (legend, *values) = (c.value for c in row)
        fields = [ct(values[idx] if values[idx] is not None else "") for idx, ct in enumerate(cls.FIELD_TYPES)]
        return cls(idx, fields)

    def to_spreadsheet(self, ws, row_number):
        ws.cell(row=row_number, column=1, value=self.legend.strip())
        for idx, field in enumerate(self.fields):
            value = str(field)
            if value is None:
                raise Exception('value is required')
            ws.cell(row=row_number, column=idx + 2, value=value)

    def to_realearn_dict(self):
        id = '%s. %s' % (self.index, self.legend)
        control = {
          "id": id,
          "name": self.label,
          "groupId": self.section,
          "source": {
            **REALEARN_CONTROL_SOURCE_COMMON,
            "character": 1 if self.physical == 'Button' else 0,
            "channel": int(str(self['Ch'])),
            "number": int(str(self['CC'])),
          },
          "mode": {
            "maxStepSize": 0.05,
            "minStepFactor": 1,
            "maxStepFactor": 5,
            "takeoverMode": "pickup-tolerant"
          },
          "target": {
                **REALEARN_CONTROL_TARGET_COMMON,
                "controlElementType": "Button" if self.physical == 'Button' else None,
                "controlElementIndex": self.label,
          },
          "feedbackIsEnabled": False,
          "visibleInProjection": False
        }
        return control

class Template():
    LINE_SIZE = 52
    MESSAGE_START=b'\xf0\x00 )\x02\x00\x7f\x00\x00'
    MESSAGE_END=b'\x124\xf7'
    FIELD_TYPES = [
        define_field(SelectValue, name='N/A 1', valid_values=[8, 17, 19, 24, 25]),
        define_field(SelectValue, name='N/A 2', valid_values=[1, 2, 3, 10]),
        Pad1,
        define_field(NumericValue, name='N/A 3'),
        define_field(TemplateName, name='Name'),
        define_field(NumericValue, name='N/A 4'),
        define_field(ManufacturerName, name='Manufacturer'),
        define_field(SelectValue, name='Channel', valid_values=[0, 16]),
        define_field(SelectValue, name='Midi port | Keyb MIDI Chan', valid_values=[0, 16, 48, 53, 54, 112]),
        Pad2,
        define_field(SelectValue, name='Velocity curve', valid_values=[0, 1, 2, 3]),
        define_field(SelectValue, name='Select 2', valid_values=[4, 5]),
        define_field(SelectValue, name='Aftertouch | Auto Snapshot | Not Synth', valid_values=[0, 2, 3, 4, 5, 6, 7]),
        define_field(NumericValue, name='Override MIDI Ch'),
        define_field(SelectValue, name='Touchpad X Type', valid_values=[0, 1, 2]),
        define_field(SelectValue, name='Touchpad Y Type', valid_values=[0, 1, 2]),
        Pad2,
        define_field(SelectValue, name='Select 6', valid_values=[48, 64, 79]),
        define_field(SelectValue, name='Select 7', valid_values=[0, 64]),
        Pad2,
        define_field(SelectValue, name='Select 8', valid_values=[64, 78]),
        define_field(SelectValue, name='Select 9', valid_values=[64, 127]),
        Pad2,
        define_field(SelectValue, name='Select 10', valid_values=[64]),
        define_field(SelectValue, name='Select 11', valid_values=[127]),
        Pad3,
        define_field(SelectValue, name='Select 12', valid_values=[0,1]),
        define_field(SelectValue, name='Select 13', valid_values=[0,49]),
        define_field(SelectValue, name='Select 14', valid_values=[64,68]),
        define_field(SelectValue, name='Select 15', valid_values=[43,64]),
        define_field(SelectValue, name='Select 16', valid_values=[127]),
        Pad6,
        define_field(SelectValue, name='Select 17', valid_values=[0,18]),
        define_field(SelectValue, name='Select 18', valid_values=[3,5]),
        define_field(SelectValue, name='Select 19', valid_values=[90]),
        define_field(SelectValue, name='Select 19', valid_values=[0,64]),
        Pad7,
        define_field(SelectValue, name='Select 19', valid_values=[0,1]),
        define_field(SelectValue, name='Select 19', valid_values=[20]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad5,
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 32]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad3,
        define_field(SelectValue, name='Select 19', valid_values=[100]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        Pad3,
        define_field(SelectValue, name='Select 19', valid_values=[52, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[64, 78]),
        define_field(SelectValue, name='Select 19', valid_values=[64, 72]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad6,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad4,
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[127]),
        Pad2,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64, 78]),
        Pad4,
        define_field(SelectValue, name='Select 19', valid_values=[3, 4]),
        define_field(SelectValue, name='Select 19', valid_values=[90]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad8,
        define_field(SelectValue, name='Select 19', valid_values=[20]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64, 74]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad2,
        define_field(SelectValue, name='Select 19', valid_values=[0, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[6, 44, 64]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[64, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0,32]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[96, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad2,
        define_field(SelectValue, name='Select 19', valid_values=[0, 36]),
        define_field(SelectValue, name='Select 19', valid_values=[100]),
        define_field(SelectValue, name='Select 19', valid_values=[64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64, 65]),
        Pad2,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64, 98]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64, 74]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad6,
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        Pad3,
        define_field(SelectValue, name='Select 19', valid_values=[0, 1]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 1]),
        define_field(SelectValue, name='Midi port', valid_values=[0, 112]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 2]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 56]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 60, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 7]),
        Pad4,
        define_field(SelectValue, name='Midi port', valid_values=[0, 112]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 2]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 69]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 72, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 7]),
        Pad5,
        define_field(SelectValue, name='Select 19', valid_values=[0, 2]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[0, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 7]),
        Pad5,
        define_field(SelectValue, name='Select 19', valid_values=[0, 2]),
        Pad1,
        define_field(SelectValue, name='Select 19', valid_values=[0, 127]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 64]),
        define_field(SelectValue, name='Select 19', valid_values=[0, 7]),
        Pad170,
    ]

    def __init__(self, header_fields, controls):
        self.header_fields = header_fields
        self.controls = controls

    @property
    def bytes(self):
        bytes = bytearray(Template.MESSAGE_START)
        for field in self.header_fields:
            bytes.extend(field.bytes)
        for control in self.controls:
            bytes.extend(control.bytes)
        bytes.extend(Template.MESSAGE_END)
        return bytes

    def __getitem__(self, key):
        field = next((x for x in self.header_fields if x.name == key), None)
        if field is None:
            raise KeyError
        return field

    def __str__(self):
       return '\n'.join([str(x) for x in self.controls])

    def __len__(self):
        return len(self.bytes)

    def __eq__(self, other):
        if not isinstance(other, Template):
            # don't attempt to compare against unrelated types
            return NotImplemented

        if len(self) != len(other):
            return False

        for idx, byte in enumerate(self.bytes):
            if byte != other.bytes[idx]:
                print('DIFFERENCE AT BYTE', idx)
                print(self.bytes[:50])
                print(other.bytes[:50])
                return False
        return True

    @property
    def name(self):
        name = next(str(x) for x in self.header_fields if x.name == 'Name')
        manufacturer = next(str(x) for x in self.header_fields if x.name == 'Manufacturer')
        return ('%s %s' % (manufacturer.strip(), name.strip())).strip()

    def print_all(self, only_unknown=False):
        for control in self.controls:
            if only_unknown and control.section != '':
                continue
            # print('---- %s | %s | %s'  % (control.legend, control.full_name, control.byte_index, ))
            print(control.index, control.legend, control)
            # print('')

    def print_distinct(self, fieldName):
        print(set([c[fieldName] for c in self.controls]), sys.argv[1])

    def print_fields(self, fieldName):
        for line in self.controls:
            print('%s %s %s %s' % (line.legend, line['name'], line['CC|Note'], line[fieldName]))

    @property
    def unknowns(self):
        return [x for x in self.controls if x.section == '']

    def print_unknowns(self):
        for control in self.unknowns:
            print(control.index + 2, control.csv())

    def diff_headers(self, other):
        for idx, header in enumerate(self.header_fields):
            other_header = other.header_fields[idx]
            for byte_index, byte in enumerate(header.bytes):
                if byte != other_header.bytes[byte_index]:
                    print('-', idx + 1, header.name, header,'---->', other_header)
                    break

    def compare_unknowns(self, other):
        other_unknowns = other.controls
        for idx, control in enumerate(self.controls):
            other_control = other_unknowns[idx]
            for field_index, field in enumerate(control.fields):
                if field.name not in ('Name', 'Ch', 'CC', 'Display', 'Type', 'Pot'): # 'Name', 'Low', 'High', 'Ch', 'MSBank', 'Sysex'):
                    other_field = other_control.fields[field_index]
                    if field != other_field:
                        print(control.index + 2, control.legend, field.name, str(field), '---->', str(other_field))

    def print_controls(self):
        for control in self.controls:
            print(control.csv())

    @classmethod
    def from_sysex(cls, filename):
        with open(filename, "rb") as f:
            file_contents = f.read()

        if file_contents[:len(Template.MESSAGE_START)] != Template.MESSAGE_START:
            raise Exception('bad header')

        if file_contents[-len(Template.MESSAGE_END):] != Template.MESSAGE_END:
            raise Exception('bad footer')

        body = file_contents[len(Template.MESSAGE_START):-len(Template.MESSAGE_END)]
        offset = 405 - len(Template.MESSAGE_START)
        full_header = bytearray(body[:offset])
        controls = body[len(full_header):]

        if len(full_header) != 396:
            raise Exception('bad header bytes length')

        instance = cls(
            header_fields=[
                ct._pop_from(full_header)
                for ct in cls.FIELD_TYPES
            ],
            controls = [
                SingleControl.from_sysex(idx, controls[i:i + cls.LINE_SIZE])
                for idx, i in enumerate(range(0, len(controls), cls.LINE_SIZE))
            ],
        )

        bytes = instance.bytes

        for idx, byte in enumerate(file_contents):
            if byte != bytes[idx]:
                raise Exception('bad serializing')

        return instance

    def to_sysex(self, file):
        with open(file, "wb") as f:
            f.write(self.bytes)

    @classmethod
    def from_spreadsheet(cls, filename):
        wb = load_workbook(filename=filename)
        ws = wb['Template configuration']
        header_fields = [cls.FIELD_TYPES[idx](row[1].value if row[1].value is not None else "") for idx, row in enumerate(ws.rows)]

        ws = wb['Controls']
        controls = [SingleControl.from_spreadsheet(idx - 1, row) for idx, row in enumerate(ws.rows) if idx > 0]
        return cls(header_fields, controls)


    def to_spreadsheet(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Template configuration"

        for idx, field in enumerate(self.header_fields):
            ws.cell(row=idx+1, column=1, value=field.name)
            ws.cell(row=idx+1, column=2, value=str(field))

        wb.create_sheet("Controls")
        ws = wb['Controls']
        ws.cell(row=1, column=1, value='Legend')
        for idx, field in enumerate(self.controls[0].fields):
            ws.cell(row=1, column=idx + 2, value=field.name)
        for idx, control in enumerate(self.controls):
            control.to_spreadsheet(ws, idx + 2)

        wb.save(filename)

    def to_json(self, filename):
        id = self.name
        groups = {c.section for c in self.controls if c.section != ''}
        main = {
          "kind": "Instance",
          "version": "2.16.14",
          "value": {
            "mainUnit": {
              "version": "2.16.14",
              "id": id,
              "name": "Main",
              "stayActiveWhenProjectInBackground": "OnlyIfBackgroundProjectIsRunning",
              "livesOnUpperFloor": False,
              "controlDeviceId": "0",
              "defaultGroup": {},
              "controllerGroups": [{"id": s, "name": s} for s in groups],
              "defaultControllerGroup": {},
              "controllerMappings": [c.to_realearn_dict() for c in self.controls if c.physical is not None],
              "instanceFx": {
                "address": "Focused"
              }
            },
            "additionalUnits": []
          }
        }
        with open(filename, "w") as f:
            f.write(json.dumps(main, indent=2))

parser = argparse.ArgumentParser(
    prog='x-station-sheet',
    description='Converts from Novation X-station Sysex to Excel files and back',
)

parser.add_argument('command', choices=['xlsx', 'syx', 'json'])
parser.add_argument('filename')
args = parser.parse_args()

path = Path(args.filename).absolute()

if args.command == 'xlsx':
    if path.suffix != '.syx':
        raise Exception('expecting a \'*.syx\' file as input')
    output = path.with_suffix('.xlsx')
    template = Template.from_sysex(path)
    template.to_spreadsheet(output)
elif args.command == 'syx':
    if path.suffix != '.xlsx':
        raise Exception('expecting a \'*.xlsx\' file as input')
    output = path.with_suffix('.syx')
    template = Template.from_spreadsheet(path)
    template.to_sysex(output)
elif args.command == 'json':
    if path.suffix != '.xlsx':
        raise Exception('expecting a \'*.xlsx\' file as input')
    output = path.with_suffix('.json')
    template = Template.from_spreadsheet(path)
    template.to_json(output)
else:
    raise Exception('unknown command')
