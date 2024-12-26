#!/usr/bin/env python3
import argparse
import json
from openpyxl import Workbook, load_workbook
from pathlib import Path
from fields import CONTROL_FIELDS, CONTROL_TEMPLATE_FIELDS, TEMPLATE_FIELDS

with open('x-station-indices.json', 'r') as f:
    indices = json.loads(f.read())

REALEARN_TARGET_DUMMY = {
    "type": 53,
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "useProject": False,
    "moveView": False,
    "seekPlay": False,
    "oscArgIndex": 0,
    "mouseAction": { "kind": "MoveTo", "axis": "X" },
    "takeMappingSnapshot": { "kind": "LastLoaded" }
}

REALEARN_BUTTON_DESIGN = {
    "background": { "kind": "Color" },
    "foreground": { "kind": "None" },
    "static_text": ""
}

REALEARN_CONTROL_SOURCE_COMMON = {
    "isRegistered": False,
    "is14Bit": False,
    "oscArgIndex": 0,
    "buttonIndex": 0,
    "buttonDesign": REALEARN_BUTTON_DESIGN,
}

REALEARN_CONTROL_TARGET_COMMON = {
    "category": "virtual",
    "fxAnchor": "id",
    "useSelectionGanging": False,
    "useTrackGrouping": False,
    "seekBehavior": "Immediate",
    "learnable": False,
    "mouseAction": {
      "kind": "MoveTo",
      "axis": "X"
    },
    "pollForFeedback": False,
    "takeMappingSnapshot": {
      "kind": "ById",
      "id": ""
    }
}

KNOWN_TEMPLATES = {
    "Trigger": "1,127,0,01110000,4,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Toggle": "1,0,127,01111000,4,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Momentary": "1,0,127,01110100,0,0,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Continuous": "1,0,127,01110000,0,0,0,0,0,0,000000000000000000000000000000000000,0,4",
    "ContinuousPickup": "1,0,127,01110000,9,64,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Unary": "1,63,64,01110000,0,65,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Jog": "1,63,65,01110000,0,65,0,0,0,0,000000000000000000000000000000000000,0,4",
    "Pitch": "10,0,127,01110000,0,1,0,0,0,0,000000000000000000000000000000000000,0,4",
    "-": "0,0,127,01110000,0,64,0,0,0,0,000000000000000000000000000000000000,0,4",
}

class FieldSet():
    @classmethod
    def from_values(cls, definition, values, *args, **kwargs):
        fields = [ct(values[idx] if values[idx] is not None else "") for idx, ct in enumerate(definition)]
        return cls(fields, *args, **kwargs)

    @classmethod
    def from_csv(cls, definition, string, *args, **kwargs):
        values = string.split(',')
        return cls.from_values(definition, values, *args, **kwargs)

    @classmethod
    def from_bytes(cls, definition, sysex, *args, **kwargs):
        # important to be bytearray (mutable)
        if not isinstance(sysex, bytearray):
            raise Exception('expecting a bytearray instance')
        if len(sysex) != sum(j.get_length() for j in definition):
            raise Exception('bad length')

        fields = [ct._pop_from(sysex) for ct in definition]

        if len(sysex) != 0:
            raise Exception('non parsed fields')
        return cls(fields, *args, **kwargs)

    def __init__(self, fields, name=None):
        self._name = name
        self.fields = fields

    @property
    def name(self):
        name = next((str(x) for x in self.fields if x.name == 'Name'), None)
        return name if name else self._name

    @name.setter
    def name(self, value):
        self._name = value

    @property
    def bytes(self):
        bytes = bytearray()
        for field in self.fields:
            bytes.extend(field.bytes)
        return bytes

    @property
    def csv(self):
        return ','.join([str(x) for x in self.fields])

    def __len__(self):
        return len(self.bytes)

    def __str__(self):
        return '<%s: %s>' % (self.name, self.csv)

    def __getitem__(self, key):
        field = next((x for x in self.fields if x.name == key), None)
        if field is None:
            raise KeyError
        return field

    def __eq__(self, other):
        return self.bytes == other.bytes

    def get_field_names(self):
        return [field.name for field in self.fields]

    def find_index(self, field_name):
        for idx, field in enumerate(self.fields):
            if field.name == field_name:
                return idx
        return None

    def to_spreadsheet(self, ws, row_number):
        ws.cell(row=row_number, column=1, value=self.name.strip())
        for idx, field in enumerate(self.fields):
            value = str(field)
            if value is None:
                raise Exception('value is required')
            ws.cell(row=row_number, column=idx + 2, value=value)
        ws.cell(row=row_number, column=len(self.fields) + 2, value=self.bytes.hex())

class SingleControl(FieldSet):
    @classmethod
    def from_bytes(cls, *args, **kwargs):
        return super(SingleControl, cls).from_bytes(CONTROL_FIELDS, *args, **kwargs)

    @classmethod
    def from_values(cls, *args, **kwargs):
        return super(SingleControl, cls).from_values(CONTROL_FIELDS, *args, **kwargs)

    @classmethod
    def from_spreadsheet(cls, idx, row):
        (legend, template, *values) = (c.value for c in row)
        return super(SingleControl, cls).from_values(CONTROL_FIELDS, values, index=idx)

    def __init__(self, fields, index):
        super().__init__(fields)
        self.index = index

    def get_template(self):
        values = []
        for definition in CONTROL_TEMPLATE_FIELDS:
            name = definition.get_name()
            values.append(str(self[name]))

        return FieldSet.from_values(CONTROL_TEMPLATE_FIELDS, values)

    @property
    def section(self):
        return indices[self.index]['section']

    @property
    def group(self):
        selector = '(%s)' % indices[self.index]['selector'] if 'selector' in indices[self.index] else ''
        group = '%s%s' % (indices[self.index]['section'], selector)
        return group

    @property
    def label(self):
        if 'selector' not in indices[self.index]:
            return indices[self.index]['legend']
        return '%s (%s)' % (indices[self.index]['legend'], indices[self.index]['selector'])

    @property
    def physical(self):
        return indices[self.index].get('physical', None)

    @property
    def legend(self):
        legend = '%s>%s' % (self.group, indices[self.index]['legend'])
        return legend

    def to_spreadsheet(self, ws, row_number, template_name):
        ws.cell(row=row_number, column=1, value=self.legend.strip())
        ws.cell(row=row_number, column=2, value=template_name)

        template = self.get_template()

        for idx, field in enumerate(self.fields):
            pos = template.find_index(field.name)
            value = '=IFERROR(VLOOKUP($B%s,Templates!$A$2:$P$1001,%s,0), "")' % (row_number, pos + 2) if pos is not None else str(field)
            if value is None:
                raise Exception('value is required')
            ws.cell(row=row_number, column=idx + 3, value=value)

    def to_realearn_dict(self):
        id = '%s. %s' % (self.index, self.legend)
        control = {
          "id": id,
          "name": self.label,
          "groupId": self.section,
          "source": {
            **REALEARN_CONTROL_SOURCE_COMMON,
            "character": 1 if self.physical == 'Button' else 0,
            "channel": int(str(self['Ch'])),
            "number": int(str(self['CC'])),
          },
          "mode": {
            "maxStepSize": 0.05,
            "minStepFactor": 1,
            "maxStepFactor": 5,
            "takeoverMode": "pickup-tolerant"
          },
          "target": {
                **REALEARN_CONTROL_TARGET_COMMON,
                "controlElementType": "Button" if self.physical == 'Button' else None,
                "controlElementIndex": self.label,
          },
          "feedbackIsEnabled": False,
          "visibleInProjection": False
        }
        return control

class Template():
    LINE_SIZE = 52
    MESSAGE_START=b'\xf0\x00 )\x02\x00\x7f\x00\x00'
    MESSAGE_END=b'\x124\xf7'

    def __init__(self, header_fields, controls):
        self.header_fields = header_fields
        self.controls = controls

    @property
    def bytes(self):
        bytes = bytearray(Template.MESSAGE_START)
        for field in self.header_fields:
            bytes.extend(field.bytes)
        for control in self.controls:
            bytes.extend(control.bytes)
        bytes.extend(Template.MESSAGE_END)
        return bytes

    def __getitem__(self, key):
        field = next((x for x in self.header_fields if x.name == key), None)
        if field is None:
            raise KeyError
        return field

    def __str__(self):
       return '\n'.join([str(x) for x in self.controls])

    def __len__(self):
        return len(self.bytes)

    def __eq__(self, other):
        if not isinstance(other, Template):
            # don't attempt to compare against unrelated types
            return NotImplemented

        if len(self) != len(other):
            return False

        for idx, byte in enumerate(self.bytes):
            if byte != other.bytes[idx]:
                return False
        return True

    @property
    def name(self):
        name = next(str(x) for x in self.header_fields if x.name == 'Name')
        manufacturer = next(str(x) for x in self.header_fields if x.name == 'Manufacturer')
        return ('%s %s' % (manufacturer.strip(), name.strip())).strip()

    @property
    def unknowns(self):
        return [x for x in self.controls if x.section == '']

    def diff_headers(self, other):
        for idx, header in enumerate(self.header_fields):
            other_header = other.header_fields[idx]
            for byte_index, byte in enumerate(header.bytes):
                if byte != other_header.bytes[byte_index]:
                    print('-', idx + 1, header.name, header,'---->', other_header)
                    break

    def compare_unknowns(self, other):
        other_unknowns = other.controls
        for idx, control in enumerate(self.controls):
            other_control = other_unknowns[idx]
            for field_index, field in enumerate(control.fields):
                if field.name not in ('Name', 'Ch', 'CC', 'Display', 'Type', 'Pot'): # 'Name', 'Low', 'High', 'Ch', 'MSBank', 'Sysex'):
                    other_field = other_control.fields[field_index]
                    if field != other_field:
                        print(control.index + 2, control.legend, field.name, str(field), '---->', str(other_field))

    @classmethod
    def from_sysex(cls, filename):
        with open(filename, "rb") as f:
            file_contents = f.read()

        if file_contents[:len(Template.MESSAGE_START)] != Template.MESSAGE_START:
            raise Exception('bad header')

        if file_contents[-len(Template.MESSAGE_END):] != Template.MESSAGE_END:
            raise Exception('bad footer')

        body = file_contents[len(Template.MESSAGE_START):-len(Template.MESSAGE_END)]
        offset = 405 - len(Template.MESSAGE_START)
        full_header = bytearray(body[:offset])
        controls = body[len(full_header):]

        if len(full_header) != 396:
            raise Exception('bad header bytes length')

        instance = cls(
            header_fields=[
                ct._pop_from(full_header)
                for ct in TEMPLATE_FIELDS
            ],
            controls = [
                SingleControl.from_bytes(bytearray(controls[i:i + cls.LINE_SIZE]), index=idx)
                for idx, i in enumerate(range(0, len(controls), cls.LINE_SIZE))
            ],
        )

        bytes = instance.bytes

        for idx, byte in enumerate(file_contents):
            if byte != bytes[idx]:
                raise Exception('bad serializing')

        return instance

    def to_sysex(self, file):
        with open(file, "wb") as f:
            f.write(self.bytes)

    @classmethod
    def from_spreadsheet(cls, filename):
        wb = load_workbook(filename=filename, data_only=True)
        ws = wb['Template configuration']
        header_fields = [TEMPLATE_FIELDS[idx](row[1].value if row[1].value is not None else "") for idx, row in enumerate(ws.rows)]

        ws = wb['Controls']
        controls = [SingleControl.from_spreadsheet(idx - 1, row) for idx, row in enumerate(ws.rows) if idx > 0]
        return cls(header_fields, controls)

    def __get_control_templates(self):
        permutations = {bytes(t.get_template().bytes) for t in self.controls}
        known = [FieldSet.from_csv(CONTROL_TEMPLATE_FIELDS, csv, name=name) for name, csv in KNOWN_TEMPLATES.items()]
        templates = []
        for idx, x in enumerate(permutations):
            name = 'template%s' % idx
            template = FieldSet.from_bytes(CONTROL_TEMPLATE_FIELDS, bytearray(x), name)
            match = next((x for x in known if x == template), None)
            if match:
                template.name = match.name
            templates.append(template)
        return templates

    def __templates_to_spreadsheet(self, wb):
        wb.create_sheet("Templates")
        wst = wb['Templates']
        for idx, field in enumerate(self.controls[0].get_template().fields):
            wst.cell(row=1, column=idx + 2, value=field.name)
        wst.cell(row=1, column=len(self.controls[0].get_template().fields) + 2, value='Bytes')
        templates = self.__get_control_templates()
        for idx, template in enumerate(templates):
            template.to_spreadsheet(wst, idx + 2)
            print(template)
        return templates

    def to_spreadsheet(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Template configuration"

        for idx, field in enumerate(self.header_fields):
            ws.cell(row=idx+1, column=1, value=field.name)
            ws.cell(row=idx+1, column=2, value=str(field))

        templates = self.__templates_to_spreadsheet(wb)

        wb.create_sheet("Controls")
        ws = wb['Controls']
        ws.cell(row=1, column=1, value='Legend')
        ws.cell(row=1, column=2, value='Template')
        for idx, name in enumerate(self.controls[0].get_field_names()):
            ws.cell(row=1, column=idx + 3, value=name)

        for idx, control in enumerate(self.controls):
            found_template = None
            for template in templates:
                if control.get_template() == template:
                    found_template = template

            if not found_template:
                raise Exception('unknown template')

            control.to_spreadsheet(ws, idx + 2, found_template.name)

        wb.save(filename)

    def to_json(self, filename):
        id = self.name
        groups = {c.section for c in self.controls if c.section != ''}
        main = {
          "kind": "Instance",
          "version": "2.16.14",
          "value": {
            "mainUnit": {
              "version": "2.16.14",
              "id": id,
              "name": "Main",
              "stayActiveWhenProjectInBackground": "OnlyIfBackgroundProjectIsRunning",
              "livesOnUpperFloor": False,
              "controlDeviceId": "0",
              "defaultGroup": {},
              "controllerGroups": [{"id": s, "name": s} for s in groups],
              "defaultControllerGroup": {},
              "controllerMappings": [c.to_realearn_dict() for c in self.controls if c.physical is not None],
              "instanceFx": {
                "address": "Focused"
              }
            },
            "additionalUnits": []
          }
        }
        with open(filename, "w") as f:
            f.write(json.dumps(main, indent=2))

parser = argparse.ArgumentParser(
    prog='x-station-sheet',
    description='Converts from Novation X-station Sysex to Excel files and back',
)

parser.add_argument('command', choices=['xlsx', 'syx', 'json'])
parser.add_argument('filename')
parser.add_argument("--output", help="output")
args = parser.parse_args()

path = Path(args.filename).absolute()

if args.command == 'xlsx':
    if path.suffix != '.syx':
        raise Exception('expecting a \'*.syx\' file as input')
    output = args.output if args.output else path.with_suffix('.xlsx')
    template = Template.from_sysex(path)
    template.to_spreadsheet(output)
elif args.command == 'syx':
    if path.suffix != '.xlsx':
        raise Exception('expecting a \'*.xlsx\' file as input')
    output = args.output if args.output else path.with_suffix('.syx')
    template = Template.from_spreadsheet(path)
    template.to_sysex(output)
elif args.command == 'json':
    if path.suffix != '.xlsx':
        raise Exception('expecting a \'*.xlsx\' file as input')
    output = args.output if args.output else path.with_suffix('.json')
    template = Template.from_spreadsheet(path)
    template.to_json(output)
else:
    raise Exception('unknown command')
